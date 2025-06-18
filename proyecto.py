import sys
import os
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                            QLabel, QLineEdit, QPushButton, QProgressBar, QTextEdit, 
                            QFileDialog, QMessageBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

class ExcelWorker(QThread):
    progress_changed = pyqtSignal(int)
    message_logged = pyqtSignal(str)
    finished = pyqtSignal(str)
    error_occurred = pyqtSignal(str)

    def __init__(self, input_file, output_file):
        super().__init__()
        self.input_file = input_file
        self.output_file = output_file

    def run(self):
        try:
            self.message_logged.emit("Cargando archivo de entrada...")
            self.progress_changed.emit(10)
            
            # Cargar archivo de entrada
            wb = load_workbook(self.input_file)
            sheet = wb.active
            
            # Crear archivo de salida
            wb_result = load_workbook(self.input_file)
            if "Resultados" in wb_result.sheetnames:
                wb_result.remove(wb_result["Resultados"])
            ws_result = wb_result.create_sheet("Resultados")
            
            # Configurar encabezados
            headers = ["Función", "Acción", "Código Asignado"]
            for col, header in enumerate(headers, 1):
                cell = ws_result.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            self.progress_changed.emit(30)
            
            current_row = 2
            total_rows = sheet.max_row
            
            # Procesar cada fila
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                if len(row) < 2:  # Verificar que tenga al menos 2 columnas
                    continue
                
                nombre = str(row[0]) if row[0] else ""
                accion = str(row[1]).strip().lower() if row[1] else ""
                
                # Solo procesar si dice "mantener"
                if "mantener" in accion:
                    # Asignar código basado en el nombre
                    codigo = self.generar_codigo(nombre)
                    
                    # Escribir resultados
                    ws_result.cell(row=current_row, column=1, value=nombre)
                    ws_result.cell(row=current_row, column=2, value=row[1] if len(row) > 1 else "")
                    ws_result.cell(row=current_row, column=3, value=codigo)
                    
                    current_row += 1
                
                # Actualizar progreso
                progress = 30 + (row_idx / total_rows * 65)
                self.progress_changed.emit(int(progress))
            
            # Ajustar formato
            ws_result.column_dimensions['A'].width = 50
            ws_result.column_dimensions['B'].width = 15
            ws_result.column_dimensions['C'].width = 40
            
            # Eliminar hoja original si existe
            if "Sheet" in wb_result.sheetnames:
                wb_result.remove(wb_result["Sheet"])
            
            # Guardar resultados
            wb_result.save(self.output_file)
            self.progress_changed.emit(100)
            self.message_logged.emit(f"Proceso completado. {current_row-2} registros procesados.")
            self.finished.emit(self.output_file)
            
        except Exception as e:
            self.error_occurred.emit(f"Error en fila {row_idx}: {str(e)}")

    def generar_codigo(self, nombre):
        """Genera un código T24 estándar basado en el nombre de la función"""
        if not nombre:
            return "CODIGO_NO_ASIGNADO"
        
        # Simplificar el nombre para crear el código
        codigo = "ENQ." + nombre.upper() \
            .replace(" ", ".") \
            .replace("Á", "A") \
            .replace("É", "E") \
            .replace("Í", "I") \
            .replace("Ó", "O") \
            .replace("Ú", "U") \
            .replace("Ñ", "N") \
            .replace("(", "") \
            .replace(")", "") \
            .replace(",", "")
        
        return codigo[:50]  # Limitar longitud

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Generador de Códigos T24")
        self.setGeometry(100, 100, 900, 700)
        self.worker = None
        self.init_ui()

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Título
        title = QLabel("Generador de Códigos para Perfiles T24")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #005a9c;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # Sección de archivo de entrada
        input_layout = QHBoxLayout()
        input_label = QLabel("Archivo de entrada:")
        input_label.setStyleSheet("font-size: 18px;")
        self.input_edit = QLineEdit()
        self.input_edit.setPlaceholderText("Seleccione el archivo Excel...")
        input_button = QPushButton("Examinar")
        input_button.clicked.connect(self.select_input_file)
        
        input_layout.addWidget(input_label)
        input_layout.addWidget(self.input_edit, 1)
        input_layout.addWidget(input_button)
        layout.addLayout(input_layout)

        # Sección de archivo de salida
        output_layout = QHBoxLayout()
        output_label = QLabel("Archivo de salida:")
        output_label.setStyleSheet("font-size: 18px;")
        self.output_edit = QLineEdit()
        self.output_edit.setPlaceholderText("Especifique el archivo de resultados...")
        output_button = QPushButton("Examinar")
        output_button.clicked.connect(self.select_output_file)
        
        output_layout.addWidget(output_label)
        output_layout.addWidget(self.output_edit, 1)
        output_layout.addWidget(output_button)
        layout.addLayout(output_layout)

        # Botón de procesamiento
        self.process_button = QPushButton("Generar Códigos")
        self.process_button.setStyleSheet("""
            QPushButton {
                font-size: 18px;
                padding: 10px 20px;
                background-color: #27ae60;
                color: white;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
            QPushButton:disabled {
                background-color: #95a5a6;
            }
        """)
        self.process_button.clicked.connect(self.process_file)
        layout.addWidget(self.process_button, 0, Qt.AlignHCenter)

        # Barra de progreso
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        layout.addWidget(self.progress_bar)

        # Área de registro
        log_label = QLabel("Registro de actividad:")
        log_label.setStyleSheet("font-weight: bold;")
        layout.addWidget(log_label)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("font-family: Consolas; font-size: 18px;")
        layout.addWidget(self.log_text, 1)

    def select_input_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar archivo Excel", "", 
            "Archivos Excel (*.xlsx *.xls);;Todos los archivos (*)"
        )
        if file_path:
            self.input_edit.setText(file_path)
            dir_name = os.path.dirname(file_path)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            self.output_edit.setText(os.path.join(dir_name, f"{base_name}_CODIGOS.xlsx"))
            self.log_text.append(f"Archivo de entrada seleccionado: {file_path}")

    def select_output_file(self):
        default_path = self.output_edit.text() or os.path.expanduser("~")
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Guardar resultados", default_path,
            "Archivos Excel (*.xlsx);;Todos los archivos (*)"
        )
        if file_path:
            if not file_path.lower().endswith('.xlsx'):
                file_path += '.xlsx'
            self.output_edit.setText(file_path)
            self.log_text.append(f"Archivo de salida seleccionado: {file_path}")

    def process_file(self):
        input_file = self.input_edit.text().strip()
        output_file = self.output_edit.text().strip()

        if not input_file:
            QMessageBox.warning(self, "Error", "Seleccione un archivo de entrada")
            return
            
        if not output_file:
            QMessageBox.warning(self, "Error", "Especifique un archivo de salida")
            return
            
        if not os.path.exists(input_file):
            QMessageBox.critical(self, "Error", f"El archivo no existe:\n{input_file}")
            return

        self.log_text.clear()
        self.log_text.append("Iniciando generación de códigos...")
        self.progress_bar.setValue(0)
        self.process_button.setEnabled(False)

        self.worker = ExcelWorker(input_file, output_file)
        self.worker.progress_changed.connect(self.progress_bar.setValue)
        self.worker.message_logged.connect(self.log_text.append)
        self.worker.finished.connect(self.on_process_finished)
        self.worker.error_occurred.connect(self.on_process_error)
        self.worker.start()

    def on_process_finished(self, output_file):
        self.process_button.setEnabled(True)
        QMessageBox.information(
            self, 
            "Proceso completado", 
            f"Archivo generado correctamente:\n{output_file}"
        )
        self.log_text.append("Proceso finalizado exitosamente")

    def on_process_error(self, error_msg):
        self.process_button.setEnabled(True)
        QMessageBox.critical(self, "Error", error_msg)
        self.log_text.append(f"Error: {error_msg}")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())